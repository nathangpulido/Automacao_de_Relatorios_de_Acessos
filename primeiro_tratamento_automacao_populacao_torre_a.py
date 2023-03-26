import openpyxl as opxl
from openpyxl.styles import Font, Fill, PatternFill, Alignment

# abre planilha excel
workbook = opxl.load_workbook('ocupacao_torrea_diaria.xlsx')

# seleciona a "folha" atual
sheet = workbook.active

# excluí a coluna A e as duas primeiras linhas
sheet.delete_cols(1, 1)
sheet.delete_rows(1, 2)

# separa os valores das células por \
for cell in sheet['A']:
    if cell.value is not None:
        partes = cell.value.split("\\")
    print(partes)
    for i, informação in enumerate(partes):
       if i >= 4:
            sheet.cell(row=cell.row, column=i).value += "_" + informação
       else:
        sheet.cell(row=cell.row, column=i+1).value = informação

# excluí as colunas root e crachas
sheet.delete_cols(1, 2)

# cria um dicionario para armazenar empresa e colaborador
empresa_colaborador = []
for linha in sheet.iter_rows(min_row=1, values_only=True):
    empresa_colaborador.append([linha[0], linha[1]])

# organiza a lista em ordem alfabetica
empresa_colaborador_ordenada = sorted(empresa_colaborador, key=lambda x: (x[0],x[1]))

# remove duplicatas
sem_duplicatas = []

for lista in empresa_colaborador_ordenada:
    if lista not in sem_duplicatas:
        sem_duplicatas.append(lista)
print(sem_duplicatas)

# remove terceiros e visitantes
palavras = ['ALAMO', 'CBRE', 'SMART', 'VERZANI', 'VISITANTE', 'PROVISORIO','VISIT']

sem_terceiros = sem_duplicatas.copy()  # Cria uma cópia da lista de sem_duplicatas

for sublista in sem_duplicatas:
    if any(palavra in sublista[0] or palavra in sublista[1] for palavra in palavras):
        sem_terceiros.remove(sublista)

# percorrer todas as células da planilha e atribuir valor nulo
for row in sheet.iter_rows():
    for cell in row:
        cell.value = None

# formatação das células do cabeçalho
fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
font = Font(color="FFFFFF", bold=True)
alignment = Alignment(horizontal="center", vertical="center")

# cria as colunas empresa colaborador e andar
cabeçalho1 = ['EMPRESA', 'COLABORADOR', 'ANDAR']
coluna_cabeçalho1 = 1
for item in cabeçalho1:
    cell = sheet.cell(row=1, column=coluna_cabeçalho1, value=item)
    cell.fill = fill
    cell.font = font
    cell.alignment = alignment
    coluna_cabeçalho1 +=1

# escreve na planilha a empresa e seu respectivo funcionario em ordem alfabetica
for row, sublista in enumerate(sem_terceiros, start=2):
    sheet.cell(row=row, column=1).value = sublista[0]  # escrever na coluna A
    sheet.cell(row=row, column=2).value = sublista[1]  # escrever na coluna B

# cria tabela do quantitativo de colaboradores por andar
cabeçalho2 = ['ANDAR', 'EMPRESA', 'QTDE']
coluna_cabeçalho2 = 6
for item in cabeçalho2:
    cell = sheet.cell(row=3, column=coluna_cabeçalho2, value=item)
    cell.fill = fill
    cell.font = font
    cell.alignment = alignment
    coluna_cabeçalho2 +=1

meu_dict = {'VAGO':['6°','7°','8°','9°','10°','11°'], 'KPMG':['12°'], 'EVONIK':['13°','14°'],'VAGO2':['15°','16°','17°','18°'],'OI':['19°'],
            'VAGO3':['20°'], 'HUAWEI':['21°','22°','23°','24°','25°'],'GWM':['26°'],
            'SERVICE IT':['27°'], 'NIVEA':['27°','28°'], 'HUAWEI2':['29°'],'MERCURIUS': ['29°'], 'BEIGENE': ['29°'], 'TRW':['30°','31°']}
row = 4
for key, values in meu_dict.items():
    for value in values:
        sheet.cell(row=row, column=6, value = value).alignment = Alignment(horizontal='left')
        sheet.cell(row=row, column=7, value = key).alignment = Alignment(horizontal='left')
        row += 1

# salva alterações na planilha
workbook.save('ocupacao_torrea_diaria_1°.xlsx')
