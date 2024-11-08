import pyodbc
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# Define a função que irá realizar a consulta e retornar o resultado
def execute_query(connection, sublist):
    cursor = connection.cursor()
    query = "SELECT Department, UIname, info2  FROM Personnel WHERE Department = ? AND UIname = ?"
    cursor.execute(query, (sublist[0], sublist[1]))
    result = cursor.fetchone()
    return result

# Define os parâmetros de conexão com o banco de dados
server = 'DESKTOP-2L34EIR\SQLEXPRESS'
database = 'ContinuumDB'
username = 'xxxx'
password = 'xxxxx'

# Realiza a tentativa de estabelecer a conexão com o banco de dados
try:
    connection = pyodbc.connect('DRIVER={SQL Server};SERVER=' + server + ';DATABASE=' + database + ';UID=' + username + ';PWD=' + password)
except pyodbc.Error as e:
    print(f'Erro ao conectar ao banco de dados: {e}')
else:
    # Se a conexão foi estabelecida com sucesso, imprime a mensagem de sucesso e continua com o restante do código
    print('Conexão estabelecida com sucesso!')

# Carrega a planilha
workbook = load_workbook(filename='exemplo.xlsx')

# Seleciona a folha atual da planilha
sheet = workbook.active

# Cria uma lista vazia
lista_de_funcionarios_a_atualizar = []

# Itera pelas linhas da coluna C
for row_number, row in enumerate(sheet.iter_rows(min_row=2, min_col=3, max_col=3, values_only=True), start=2):
    # Verifica se o valor da célula é igual a "?"
    if row[0] == "?":
         # Armazena os valores das células A e B na mesma linha na lista funcionarios_a_atualizar
        funcionarios_a_atualizar = [sheet.cell(row=row_number, column=1).value,
                                        sheet.cell(row=row_number, column=2).value]
        # Adiciona funcionarios_a_atualizar a lista_de_funcionarios_a_atualizar
        lista_de_funcionarios_a_atualizar.append(funcionarios_a_atualizar)

# Imprime a lista com as sublistas armazenadas
print(lista_de_funcionarios_a_atualizar)

# Itera sobre as sublistas, constrói a cláusula WHERE e executa a consulta
list_of_results = []
for sublist in lista_de_funcionarios_a_atualizar:
    result = execute_query(connection, sublist)
    list_of_results.append(result)

# Fecha a conexão com o banco de dados
connection.close()

# Imprime a lista de resultados
print(list_of_results)

# abre o arquivo FUNCIONÁRIOS GERAIS TORRE A e seleciona a 'folha' ativa
workbook1 = load_workbook(filename='FUNCIONÁRIOS GERAIS TORRE A.xlsx')
sheet1 = workbook1.active

# identifica a última linha com um valor não nulo na coluna A de FUNCIONÁRIOS GERAIS TORRE A
last_row = 1
for row in sheet1.iter_rows():
    if row[0].value is not None:
        last_row = row[0].row

# Escreve os valores do índice 1 e 2 de cada tupla(resultados das consultas ao ContinuumDb)  após a última linha não nula
for tupla in list_of_results:
    sheet1.cell(row=last_row+1, column=1, value=tupla[1]).alignment = Alignment(horizontal='left')
    sheet1.cell(row=last_row+1, column=2, value=tupla[2]).alignment = Alignment(horizontal='left')
    last_row += 1

# salva as alterações no arquivo do Excel
workbook1.save('FUNCIONÁRIOS GERAIS TORRE A.xlsx')
