from openpyxl import load_workbook
from openpyxl.styles import Alignment

# Carrega a planilha
workbook1 = load_workbook(filename='FUNCIONÁRIOS GERAIS TORRE A.xlsx')

# Seleciona a folha atual da planilha
sheet1 = workbook1.active

# Cria uma lista vazia
lista2 = []

# Itera sobre as linhas da planilha
for row in sheet1.iter_rows(min_row=1, values_only=True):

    # Cria um dicionário com os valores das colunas "A" e "B" da linha atual
    dicionario = {row[0]: row[1]}

    # Adiciona o dicionário à lista2
    lista2.append(dicionario)

# Exibe a lista
print(lista2)


# Carrega a planilha
workbook2 = load_workbook(filename='ocupacao_torrea_diaria_1°.xlsx')

# Seleciona a folha da planilha
sheet2 = workbook2.active

# Cria uma lista vazia
lista1 = []

# Itera sobre as linhas da planilha, começando pela segunda linha (ignorando a primeira linha de cabeçalho)
for row in sheet2.iter_rows(min_row=2, values_only=True):

    # Verifica se a coluna C da linha atual é nula
    if row[2] is None:
        # Cria um dicionário com os valores das colunas "B" e "C" da linha atual
        dicionario = {row[1]: '?'}

    # Adiciona o dicionário à lista1
    lista1.append(dicionario)

# Exibe a lista
print(lista1)


# Loop para iterar sobre cada dicionário da lista1
for d1 in lista1:
    # Loop para iterar sobre cada dicionário da lista2
    for d2 in lista2:
        # Verifica se há alguma chave em comum entre os dois dicionários
        if set(d1.keys()) & set(d2.keys()):
            # Se houver, atribui o valor correspondente do dicionário da lista2 ao dicionário da lista1
            for chave in set(d1.keys()) & set(d2.keys()):
                d1[chave] = d2[chave]

# Exibe a lista1 com os valores atualizados
print(lista1)

# Itera sobre a coluna B
for cell in sheet2['B']:
    # Verifica se o valor da célula é igual a alguma chave dos dicionários da lista
    for d in lista1:
        if cell.value in d.keys():
            # Se for igual, atribui o valor correspondente a essa chave à célula da coluna C
            sheet2.cell(row=cell.row, column=3, value=d[cell.value]).alignment = Alignment(horizontal='left')

# Salva a planilha com as alterações
workbook2.save('exemplo.xlsx')